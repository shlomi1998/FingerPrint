
import openpyxl
from mysql import *

fingerprints = get_all_data_fingerprint()
print(fingerprints)

fingerprints_wb = openpyxl.Workbook()
fingerprints_xlsx = fingerprints_wb.active

for entry in fingerprints:
    index, IdUser, Hand, Finger, Date, Time, LinkImage, Details, InterestingPoints = entry
    print(IdUser)
    print(Hand)
    print(Finger)
    print(Date)
    print(Time)
    print(LinkImage)
    print(Details)
    print(InterestingPoints)
    print('==================================================================================')
    fingerprints_xlsx.append(entry)

fingerprints_wb.save('fingerprints.xlsx')



users = get_all_data_users()
users_wb = openpyxl.Workbook()
users_xlsx = users_wb.active
for entry in users:
    users_xlsx.append(entry)

users_wb.save('users.xlsx')

user_relationships = get_all_data_user_relationships()
user_relationships_wb = openpyxl.Workbook()
user_relationships_xlsx = user_relationships_wb.active
for entry in user_relationships_xlsx:
    user_relationships_xlsx.append(entry)

user_relationships_wb.save('user_relationships.xlsx')


